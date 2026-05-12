"""CSV / XLSX export of `state.db`. No Migaku API needed."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from .models import CachedRow
from .supabase_client import SupabaseClient


log = logging.getLogger("migaku-supabase")


# Canonical column order. (Display name, CachedRow attribute name).
# Mirrors the Supabase table's human-facing columns.
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Word",             "dict_form"),
    ("Pinyin",           "pinyin_marks"),
    ("Meaning",          "_meaning"),         # filled from cache or --with-meaning
    ("Example",          "example"),          # v2
    ("Pinyin (numeric)", "pinyin_numeric"),
    ("Status",           "known_status"),
    ("Frequency",        "frequency_stars"),  # v2 (1-5)
    ("Fail rate %",      "fail_rate"),
    ("Total reviews",    "total_reviews"),
    ("Failed reviews",   "failed_reviews"),
    ("Part of speech",   "part_of_speech"),
    ("Language",         "lang"),
    ("First learning at", "first_learning_at"),
    ("First known at",   "first_known_at"),
    ("Last synced",      "last_synced"),
    ("Migaku key",       "migaku_key"),
    ("Sense #",          "sense_index"),
]


def _row_value(row: CachedRow, attr: str, meanings: dict[str, str] | None) -> Any:
    if attr == "_meaning":
        # Prefer the Supabase-side meaning if --with-meaning fetched it
        # because users may edit that column directly.
        # Fall back to the cache's `meaning` field, which v2 populates
        # from Migaku's published dict whenever there's a hit.
        sink_meaning = (meanings or {}).get(row.migaku_key)
        if sink_meaning:
            return sink_meaning
        return row.meaning or ""
    return getattr(row, attr, "")


def fetch_meanings_from_supabase(supabase: SupabaseClient) -> dict[str, str]:
    """One full Supabase query just for the Meaning column, keyed by Migaku key."""
    log.info("Fetching Meaning column from Supabase ...")
    records = supabase.query_all_rows()
    out: dict[str, str] = {}
    for record in records:
        key = record.get("migaku_key") or ""
        if not key:
            continue
        meaning = record.get("meaning") or ""
        if meaning:
            out[key] = meaning
    log.info("Got %d meanings (out of %d Supabase rows)", len(out), len(records))
    return out


def filter_rows(rows: list[CachedRow], lang: str | None,
                statuses: list[str] | None, include_archived: bool) -> list[CachedRow]:
    """Apply CLI filters to cache rows. Returns a stable-sorted list."""
    out = []
    status_set = set(statuses) if statuses else None
    for r in rows:
        if not include_archived and r.archived:
            continue
        if lang and r.lang != lang:
            continue
        if status_set and (r.known_status or "") not in status_set:
            continue
        out.append(r)
    out.sort(key=lambda r: r.migaku_key)
    return out


def export_csv(path: Path, rows: list[CachedRow], meanings: dict[str, str] | None) -> None:
    import csv
    headers = [name for name, _ in EXPORT_COLUMNS]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([_row_value(r, attr, meanings) for _, attr in EXPORT_COLUMNS])
    log.info("Wrote CSV: %s (%d rows)", path, len(rows))


def export_xlsx(path: Path, rows: list[CachedRow], meanings: dict[str, str] | None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "XLSX export requires `openpyxl`. Install with `pip install openpyxl` "
            "(or re-run `pip install -r requirements.txt`)."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Migaku Vocab"

    headers = [name for name, _ in EXPORT_COLUMNS]
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="EEEEEE")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill

    for r in rows:
        ws.append([_row_value(r, attr, meanings) for _, attr in EXPORT_COLUMNS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    widths = {
        "Word": 14, "Pinyin": 16, "Meaning": 50, "Example": 50,
        "Pinyin (numeric)": 18,
        "Status": 11, "Frequency": 10, "Fail rate %": 11, "Total reviews": 13,
        "Failed reviews": 14, "Part of speech": 16, "Language": 9,
        "Last synced": 22, "Migaku key": 26, "Sense #": 8,
    }
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 14)

    wb.save(path)
    log.info("Wrote XLSX: %s (%d rows)", path, len(rows))
